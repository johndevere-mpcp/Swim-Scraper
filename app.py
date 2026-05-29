"""Streamlit web UI for swim_parser.

A non-technical user uploads (or pastes) the Cal Swim workout document
and sees totals split by training group: Sprint (Dave's), Mid (Josh's),
Distance (Noah's). Whole-team sets (parser bucket 'all') are folded
into EACH group total because every group swims them — so the three
displayed columns deliberately sum to MORE than the workout total
(they overlap on whole-team yardage).

Practice-level validation flags any sub-session below MIN_PRACTICE_YD
so the user can spot either real light days (taper/recovery) or
potential parsing gaps.

Run locally:
    pip install -r requirements.txt
    streamlit run app.py

Deploy to Streamlit Community Cloud:
    1. Push this folder to a public GitHub repo (must include
       swim_parser.py, app.py, and requirements.txt).
    2. Go to https://share.streamlit.io and sign in with GitHub.
    3. Click "New app", pick the repo / branch / app.py.
    4. Click Deploy. Share the resulting URL with the team.
"""

import csv
import io
from typing import Dict, List

import streamlit as st

from swim_parser import (
    DAY_ORDER,
    WEEK_HEADER_RE,
    compute_workout_totals,
    _yard_total,
    _meter_total,
)


# ---------------------------------------------------------------------------
# Constants.
# ---------------------------------------------------------------------------

# Each displayed group total includes the whole-team ('all') bucket
# because every group swims those sets. The parser still tracks 'all'
# separately under the hood (it's needed so the four parser buckets sum
# to the workout total exactly); we only fold it in at display time.
DISPLAY_GROUP_KEYS = ["sprint", "mid", "distance"]
DISPLAY_GROUP_LABELS = {
    "sprint":   "Sprint",
    "mid":      "Mid",
    "distance": "Distance",
}
DISPLAY_GROUP_HINTS = {
    "sprint":   "Dave's group (incl. whole-team sets)",
    "mid":      "Josh's group (incl. whole-team sets)",
    "distance": "Noah's group (incl. whole-team sets)",
}

# Practice-level sanity threshold. The team's stated minimum is 2,000 yd
# per practice; sub-sessions below this are surfaced in the Validation
# section so the user can confirm they're real light days vs missed
# yardage from a parsing edge case.
MIN_PRACTICE_YD = 2000

CSV_HEADERS = [
    "Week", "Workout", "Method", "Group source",
    "Sprint yd", "Mid yd", "Distance yd",
    "Sprint m", "Mid m", "Distance m",
    "Total yd", "Total m", "Manual yd", "Checkpoint yd", "Below 2000 yd?",
]


# ---------------------------------------------------------------------------
# Formatting helpers.
# ---------------------------------------------------------------------------

def fmt_amount(yards: int, meters: int) -> str:
    """Return e.g. '12,345 yd' or '12,345 yd  /  678 m'."""
    s = f"{yards:,} yd"
    if meters:
        s += f"  /  {meters:,} m"
    return s


def group_displays(totals: Dict[str, int]) -> Dict[str, Dict[str, int]]:
    """Fold whole-team 'all' bucket into each group's displayed total.

    Returns {sprint: {y, m}, mid: {y, m}, distance: {y, m}} where each
    group's yards/meters = its own bucket + the 'all' bucket. Used by
    every display surface (banner, per-week expander, daily table,
    workouts table, CSV, Excel) so the parser can keep tracking 'all'
    separately while the UI shows the more meaningful "yards each
    group actually swims".

    The three displayed columns will sum to MORE than the workout
    total by 2*all_y; that's expected — the 'all' yardage is counted
    against every group because every group swims it.
    """
    return {
        "sprint": {
            "y": totals["sprint_y"] + totals["all_y"],
            "m": totals["sprint_m"] + totals["all_m"],
        },
        "mid": {
            "y": totals["mid_y"] + totals["all_y"],
            "m": totals["mid_m"] + totals["all_m"],
        },
        "distance": {
            "y": totals["distance_y"] + totals["all_y"],
            "m": totals["distance_m"] + totals["all_m"],
        },
    }


def render_bucket_row(totals: Dict[str, int]) -> None:
    """Render Total + 3 group metrics as a 4-column row.

    Used for both the grand-total banner and each per-week expander.
    Each group's metric includes the whole-team yardage (the parser's
    'all' bucket), so the three group columns visibly overlap on
    whole-team sets. Total stays as the workout total (each yard
    counted once).
    """
    y_total = _yard_total(totals)
    m_total = _meter_total(totals)
    disp = group_displays(totals)
    cols = st.columns(4)
    cols[0].metric("Total", fmt_amount(y_total, m_total))
    for i, key in enumerate(DISPLAY_GROUP_KEYS, start=1):
        cols[i].metric(
            DISPLAY_GROUP_LABELS[key],
            fmt_amount(disp[key]["y"], disp[key]["m"]),
            help=DISPLAY_GROUP_HINTS[key],
        )


def daily_to_markdown(daily_subtotals: Dict, week_num: int) -> str:
    """Render the per-day rollup for one week as a markdown table.

    Per the boss's spec: mini-microcycle = day, so this table sums all
    sub-sessions (AM + PM, multiple coaches) that fall on the same
    weekday into a single row. Sorted Mon → Sun via DAY_ORDER.
    Each group column folds in the whole-team 'all' yardage (see
    group_displays), so the three group columns overlap on whole-team
    sets and won't sum to the day total.
    """
    lines = [
        "| Day | Sprint | Mid | Distance | Total yd | Total m |",
        "|:---|---:|---:|---:|---:|---:|",
    ]
    keys = [(wn, d) for (wn, d) in daily_subtotals if wn == week_num]
    keys.sort(key=lambda k: DAY_ORDER.get(k[1], 999))
    for k in keys:
        dt = daily_subtotals[k]
        dy = _yard_total(dt)
        dm = _meter_total(dt)
        if dy == 0 and dm == 0:
            continue
        disp = group_displays(dt)
        lines.append(
            f"| {k[1].capitalize()} | {disp['sprint']['y']:,} "
            f"| {disp['mid']['y']:,} | {disp['distance']['y']:,} "
            f"| **{dy:,}** | {dm:,} |"
        )
    return "\n".join(lines) if len(lines) > 2 else "_No daily data for this week._"


def is_below_threshold(totals: Dict[str, int]) -> bool:
    """True if the workout has non-zero yardage strictly below MIN_PRACTICE_YD.

    Zero-yard sub-sessions are excluded — those are typically drylands
    or notes, not yardage-bearing practices, so flagging them as 'below
    minimum' would be noise. Meter-only workouts also pass through (we
    only check the yard total).
    """
    y = _yard_total(totals)
    return 0 < y < MIN_PRACTICE_YD


def attribution_label(w: dict) -> str:
    """Human-readable provenance of a sub-session's group attribution.

    Distinguishes doc-labeled splits from coach-inferred assumptions so
    the user can see when a number rests on an assumption vs an explicit
    label in the workout:
      labeled     -> 'by label'   (explicit SPRINT/MID/DISTANCE sections)
      assumed     -> 'GROUP (assumed)' e.g. 'distance (assumed)' — inferred
                     from the Coach: line, no group label in the workout
      unspecified -> 'whole-team' (no sections, no coach -> all)
    """
    attr = w.get("attribution", "")
    if attr == "labeled":
        return "by label"
    if attr == "assumed":
        return f"{w['default_group']} (assumed)"
    return "whole-team"


def workouts_to_markdown(workouts: List[dict], header_max: int = 60) -> str:
    """Render the workout list as a GitHub-flavored markdown table.

    Used instead of st.dataframe because st.dataframe pulls in pyarrow,
    which has frequent x86_64/arm64 mismatch issues on macOS. Markdown
    tables render natively in Streamlit with no extra dependency.
    Long workout headers are truncated to header_max chars so the table
    stays readable on narrow screens. A ⚠️ flag is appended to any
    workout under MIN_PRACTICE_YD so it's spottable in-table. The
    'Group source' column shows whether the split came from explicit
    doc labels or was assumed from the coach (see attribution_label).
    """
    lines = [
        "| Week | Workout | Group source | Method "
        "| Sprint | Mid | Distance | Total yd | Total m |",
        "|---:|:---|:---|:---:|---:|---:|---:|---:|---:|",
    ]
    for w in workouts:
        t = w["totals"]
        hdr = w["header"]
        if len(hdr) > header_max:
            hdr = hdr[:header_max - 1] + "…"
        # Escape pipes so a stray '|' in a header doesn't break the table.
        hdr = hdr.replace("|", "\\|")
        flag = " ⚠️" if is_below_threshold(t) else ""
        disp = group_displays(t)
        lines.append(
            f"| {w['week']} | {hdr}{flag} | {attribution_label(w)} | {w['method']} "
            f"| {disp['sprint']['y']:,} | {disp['mid']['y']:,} "
            f"| {disp['distance']['y']:,} "
            f"| {_yard_total(t):,} | {_meter_total(t):,} |"
        )
    return "\n".join(lines)


def min_week_in_text(text: str) -> int:
    """Return the smallest 'Week N' header found in text (high sentinel if none).

    Used to sort multi-file uploads chronologically — a file containing
    'Week 1..10' sorts before one containing 'Week 11..17', regardless of
    the order the user dropped them in the uploader. Files without any
    Week header fall to the end of the order.
    """
    weeks = []
    for line in text.splitlines():
        m = WEEK_HEADER_RE.match(line)
        if m:
            weeks.append(int(m.group(1)))
    return min(weeks) if weeks else 10_000


def combine_uploaded_files(uploaded_files) -> tuple:
    """Read + concatenate multiple uploaded .txt files in week order.

    Returns (combined_text, ordered_file_names). The combined text has
    each file separated by a blank line so split_workouts() can't merge
    a workout across the file boundary. Order is determined by the
    minimum Week N in each file, so 'Weeks 1-10.txt' is processed
    before 'Weeks 11-17.txt' even if dropped in the opposite order.
    """
    decoded = []
    for uf in uploaded_files:
        # utf-8-sig strips a leading BOM if present. Without this, a BOM at
        # the start of the file prefixes ﻿ onto the first line and the
        # WEEK_HEADER_RE on that line silently fails to match, causing all
        # the file's content to be tagged as the previous file's last week.
        text = uf.read().decode("utf-8-sig")
        decoded.append((uf.name, text, min_week_in_text(text)))
    decoded.sort(key=lambda t: (t[2], t[0]))
    combined = "\n\n".join(text for _, text, _ in decoded)
    ordered_names = [name for name, _, _ in decoded]
    return combined, ordered_names


def workouts_to_csv(workouts: List[dict]) -> str:
    """Render the full workout list as a CSV string for download.

    Each group column folds in the whole-team 'all' bucket (see
    group_displays); the last column flags sub-sessions below
    MIN_PRACTICE_YD so the user can filter/sort on it in Excel.
    """
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(CSV_HEADERS)
    for w in workouts:
        t = w["totals"]
        disp = group_displays(t)
        writer.writerow([
            w["week"], w["header"], w["method"], attribution_label(w),
            disp["sprint"]["y"], disp["mid"]["y"], disp["distance"]["y"],
            disp["sprint"]["m"], disp["mid"]["m"], disp["distance"]["m"],
            _yard_total(t), _meter_total(t),
            w.get("manual_total", ""), w.get("checkpoint_total", ""),
            "YES" if is_below_threshold(t) else "",
        ])
    return buf.getvalue()


def build_xlsx(results: Dict) -> bytes:
    """Render the full results as a 5-sheet Excel workbook.

    Mirrors the website's hierarchy plus a dedicated validation tab:
      Sheet 1 'Summary'      - Grand total + per-group yards
      Sheet 2 'Per Week'     - One row per week
      Sheet 3 'Per Day'      - One row per (week, day) = mini-microcycle
      Sheet 4 'Workouts'     - One row per sub-session + Below 2000 flag
      Sheet 5 'Validation'   - One row per sub-session below MIN_PRACTICE_YD

    Group columns on every sheet fold in the whole-team 'all' yardage
    (Sprint = sprint_y + all_y, etc.), so the three group columns
    overlap on whole-team sets and don't sum to the workout total.
    A note on the Summary sheet calls this out so the spreadsheet
    can be read standalone.

    Header rows are bold with a grey fill; flagged rows on the
    Workouts sheet use a pale red fill so the boss can spot them
    visually without filtering.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    flag_fill = PatternFill("solid", fgColor="FFE0E0")

    def write_header(ws, headers):
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0
                          for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = "Summary"
    gt = results["grand_total"]
    y_total = _yard_total(gt)
    m_total = _meter_total(gt)
    disp = group_displays(gt)

    write_header(ws, ["Group", "Yards", "Meters", "% of Workout Total"])
    for key, label in [("sprint", "Sprint (Dave)"),
                       ("mid", "Mid (Josh)"),
                       ("distance", "Distance (Noah)")]:
        y = disp[key]["y"]
        m = disp[key]["m"]
        pct = (y / y_total * 100) if y_total else 0
        ws.append([label, y, m, f"{pct:.1f}%"])
    ws.append(["WORKOUT TOTAL (each yd counted once)", y_total, m_total, "100.0%"])
    for cell in ws[ws.max_row]:
        cell.font = bold
    ws.append([])
    ws.append([
        "Note: each group's yardage includes whole-team sets that all "
        "groups swim, so the three group rows sum to more than the "
        "workout total. The WORKOUT TOTAL row is the season's true "
        "unique yardage (each yard counted exactly once)."
    ])
    ws.append([])
    ws.append([f"Sub-sessions parsed: {results['workouts_parsed']}"])
    if results.get("deduped"):
        ws.append([
            f"Duplicates auto-removed: {len(results['deduped'])} "
            f"({sum(d['yards'] for d in results['deduped']):,} yd)"
        ])
    flagged_count = sum(
        1 for w in results["workouts"] if is_below_threshold(w["totals"])
    )
    ws.append([
        f"Sub-sessions below {MIN_PRACTICE_YD:,} yd: {flagged_count} "
        "(see Validation sheet)"
    ])
    autosize(ws)

    # ---- Sheet 2: Per Week ----
    ws = wb.create_sheet("Per Week")
    write_header(ws, ["Week", "Sprint", "Mid", "Distance",
                      "Workout Total yd", "Workout Total m"])
    for wn in sorted(results["weekly_subtotals"]):
        wt = results["weekly_subtotals"][wn]
        d = group_displays(wt)
        ws.append([
            f"Week {wn}",
            d["sprint"]["y"], d["mid"]["y"], d["distance"]["y"],
            _yard_total(wt), _meter_total(wt),
        ])
    autosize(ws)

    # ---- Sheet 3: Per Day ----
    ws = wb.create_sheet("Per Day")
    write_header(ws, ["Week", "Day", "Sprint", "Mid", "Distance",
                      "Workout Total yd", "Workout Total m"])
    keys = sorted(
        results["daily_subtotals"].keys(),
        key=lambda k: (k[0], DAY_ORDER.get(k[1], 999)),
    )
    for (wn, day) in keys:
        dt = results["daily_subtotals"][(wn, day)]
        d = group_displays(dt)
        ws.append([
            wn, day.capitalize(),
            d["sprint"]["y"], d["mid"]["y"], d["distance"]["y"],
            _yard_total(dt), _meter_total(dt),
        ])
    autosize(ws)

    # ---- Sheet 4: Workouts ----
    # 'Group source' tells the reader whether the split came from explicit
    # doc labels or was assumed from the coach. 'Manual yd' / 'Checkpoint yd'
    # expose both signals so a reviewer can see where they diverge (the
    # session total is the larger of the two for single-group sessions, and
    # the manual sum for labeled parallel-group sessions).
    ws = wb.create_sheet("Workouts")
    write_header(ws, ["Week", "Day", "Workout", "Group source", "Method",
                      "Sprint", "Mid", "Distance",
                      "Total yd", "Total m",
                      "Manual yd", "Checkpoint yd",
                      f"Below {MIN_PRACTICE_YD} yd?"])
    for w in results["workouts"]:
        t = w["totals"]
        d = group_displays(t)
        below = is_below_threshold(t)
        ws.append([
            w["week"], w.get("day", "").capitalize(), w["header"],
            attribution_label(w), w["method"],
            d["sprint"]["y"], d["mid"]["y"], d["distance"]["y"],
            _yard_total(t), _meter_total(t),
            w.get("manual_total", ""), w.get("checkpoint_total", ""),
            "YES" if below else "",
        ])
        if below:
            for cell in ws[ws.max_row]:
                cell.fill = flag_fill
    autosize(ws)

    # ---- Sheet 5: Validation ----
    ws = wb.create_sheet("Validation")

    # 5a: zero-group weeks (attribution-failure red flag).
    ws.append(["ZERO-GROUP CHECK"])
    ws[ws.max_row][0].font = bold
    zero_hits = []
    for wn in sorted(results["weekly_subtotals"]):
        disp = group_displays(results["weekly_subtotals"][wn])
        for key in DISPLAY_GROUP_KEYS:
            if disp[key]["y"] == 0 and disp[key]["m"] == 0:
                zero_hits.append((wn, DISPLAY_GROUP_LABELS[key]))
    if zero_hits:
        for wn, label in zero_hits:
            ws.append([f"Week {wn}: {label} has 0 yd — likely attribution gap"])
            ws[ws.max_row][0].fill = flag_fill
    else:
        ws.append(["OK — every group has yardage in every week."])
    ws.append([])

    # 5b: below-threshold sub-sessions.
    ws.append([f"SUB-SESSIONS BELOW {MIN_PRACTICE_YD:,} yd"])
    ws[ws.max_row][0].font = bold
    write_header(ws, ["Week", "Day", "Workout", "Group source",
                      "Total yd", "Total m"])
    flagged = [w for w in results["workouts"] if is_below_threshold(w["totals"])]
    flagged.sort(key=lambda w: (w["week"], w["header"]))
    for w in flagged:
        t = w["totals"]
        ws.append([
            w["week"], w.get("day", "").capitalize(), w["header"],
            attribution_label(w), _yard_total(t), _meter_total(t),
        ])
    ws.append([])
    ws.append([
        f"Rows above are sub-sessions with non-zero yardage below "
        f"{MIN_PRACTICE_YD:,} yd. These are either real light days "
        "(taper / recovery / dual-meet warmup) or a sign the parser "
        "missed yardage from that block — worth a manual check."
    ])
    autosize(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Page layout.
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="Cal Swim Workout Parser",
    layout="wide",
)

st.title("Cal Swim Workout Parser")
st.caption(
    "Upload the team workout document (or paste its text) to get total "
    "yardage broken out by training group. Sprint + Mid + Distance + All "
    "sums to the workout total for every workout, every week, and the "
    "grand total."
)

# Sidebar — input controls. Filter dropdown is rendered below in a
# second sidebar block AFTER we've read the input, so it can size its
# week range to the actual max Week N found in the uploaded files
# (covers seasons of any length without hardcoding).
with st.sidebar:
    st.header("Input")
    uploaded = st.file_uploader(
        "Upload .txt file(s)",
        type=["txt"],
        accept_multiple_files=True,
        help=(
            "Drop one file or several at once (e.g. 'Weeks 1-10.txt' and "
            "'Weeks 11-17.txt'). Files are combined by the lowest Week N "
            "they contain, so upload order doesn't matter."
        ),
    )
    st.markdown("— or —")
    pasted = st.text_area(
        "Paste workout text",
        height=180,
        placeholder="Paste the document text here...",
    )

# Read input. Multi-file uploads are concatenated in chronological order
# (by the lowest Week N they contain). Pasted text is treated as a single
# block and wins only when no files are uploaded.
text = None
combined_filenames: List[str] = []
if uploaded:
    text, combined_filenames = combine_uploaded_files(uploaded)
elif pasted.strip():
    text = pasted


def detect_max_week(t: str) -> int:
    """Return the largest 'Week N' header found in text, or 17 as fallback.

    Used to size the week-filter dropdown so it matches the season length
    actually present in the upload. The 17-week fallback covers a typical
    college dual-meet season for the case where nothing's been uploaded
    yet and we still want to show a meaningful dropdown.
    """
    weeks = []
    for line in (t or "").splitlines():
        m = WEEK_HEADER_RE.match(line)
        if m:
            weeks.append(int(m.group(1)))
    return max(weeks) if weeks else 17


max_week = detect_max_week(text)

# Sidebar — filter + footer (separate from the input block above so it
# renders AFTER input has been read and the week range is known).
with st.sidebar:
    st.header("Filter")
    week_options = ["All weeks"] + [f"Week {i}" for i in range(1, max_week + 1)]
    week_choice = st.selectbox("Show one week (optional)", week_options)
    week_filter = None if week_choice == "All weeks" else int(week_choice.split()[1])

    st.markdown("---")
    st.caption(
        "**Coach → group mapping**\n\n"
        "- Dave → Sprint\n"
        "- Josh → Mid\n"
        "- Noah → Distance\n\n"
        "Whole-team sets (no group prescribed) are folded into "
        "every group total — every group swims them."
    )
    st.caption(
        f"**Practice check:** sub-sessions below "
        f"{MIN_PRACTICE_YD:,} yd are flagged ⚠️ in tables."
    )

if not text:
    st.info(
        "Upload one or more `.txt` files (or paste workout text) in the "
        "left sidebar to get started."
    )
    st.stop()

if len(combined_filenames) > 1:
    st.caption(
        "Combined "
        + str(len(combined_filenames))
        + " files in week order: "
        + ", ".join(f"`{n}`" for n in combined_filenames)
    )


def _render_dedupe_note(deduped):
    """Show an info banner if auto-dedupe skipped any sub-sessions.

    Lists each skipped workout so the user can sanity-check that the
    removals are legitimate paste artifacts and not real workouts that
    happen to share a body with an earlier one.
    """
    if not deduped:
        return
    total = sum(d["yards"] for d in deduped)
    with st.expander(
        f"Auto-dedupe removed {len(deduped)} duplicate sub-session(s) "
        f"({total:,} yd) — click to see which",
        expanded=False,
    ):
        st.markdown(
            "These had body text that exactly matched an earlier sub-session "
            "in the document (a copy-paste artifact). The first occurrence "
            "is kept; later identical copies are excluded from totals."
        )
        for d in deduped:
            st.markdown(
                f"- **Week {d['week']}**: {d['header'][:100]} — {d['yards']:,} yd"
            )


def render_weekly_validation(results: Dict) -> None:
    """Render a per-week sanity-check pane.

    For each week:
      • Show weekly total, sub-session count, average, and minimum.
      • Flag the week ⚠️ if any sub-session is under MIN_PRACTICE_YD,
        otherwise ✅.
      • If flagged, expander lists the offending sub-sessions so the
        user can confirm they're real light days vs missed yardage.

    Zero-yard sub-sessions (drylands/notes) are excluded from min and
    flag counts but included in the sub-session count for transparency.
    """
    st.subheader("Weekly Validation")

    # Zero-group check. A whole training group reading 0 for an entire
    # week is almost always an attribution failure, not reality (every
    # squad swims every week). Surface it loudly at the top so it can't
    # be missed. Uses the folded display values (own + whole-team), which
    # is what the rest of the page shows.
    zero_hits = []
    for week_num in sorted(results["weekly_subtotals"]):
        disp = group_displays(results["weekly_subtotals"][week_num])
        for key in DISPLAY_GROUP_KEYS:
            if disp[key]["y"] == 0 and disp[key]["m"] == 0:
                zero_hits.append((week_num, DISPLAY_GROUP_LABELS[key]))
    if zero_hits:
        st.error(
            "**Zero-yardage groups detected — likely an attribution gap:**\n\n"
            + "\n".join(
                f"- Week {wn}: **{label}** has 0 yd" for wn, label in zero_hits
            )
            + "\n\nEvery squad normally swims every week, so a 0 here usually "
            "means a workout's group label wasn't recognized. Worth a manual "
            "check of that week's source text."
        )
    else:
        st.success(
            "✅ Every group has yardage in every week — no zero-group gaps."
        )

    st.caption(
        f"Expected minimum: **{MIN_PRACTICE_YD:,} yd per practice**. "
        "Weeks with sub-sessions below this are flagged ⚠️ — usually "
        "real taper / recovery / dual-meet days, but worth a glance to "
        "rule out a parsing gap."
    )

    for week_num in sorted(results["weekly_subtotals"]):
        wt = results["weekly_subtotals"][week_num]
        wy = _yard_total(wt)
        week_subs = [w for w in results["workouts"] if w["week"] == week_num]
        nonzero_yards = [_yard_total(w["totals"]) for w in week_subs
                         if _yard_total(w["totals"]) > 0]
        flagged = [w for w in week_subs if is_below_threshold(w["totals"])]

        n_sub = len(week_subs)
        min_yd = min(nonzero_yards) if nonzero_yards else 0
        avg_yd = sum(nonzero_yards) // len(nonzero_yards) if nonzero_yards else 0

        if flagged:
            label = (
                f"⚠️ Week {week_num} — {wy:,} yd — "
                f"{n_sub} sub-session(s), {len(flagged)} below "
                f"{MIN_PRACTICE_YD:,} yd  (min {min_yd:,}, avg {avg_yd:,})"
            )
            with st.expander(label, expanded=False):
                st.markdown(
                    f"**{len(flagged)} sub-session(s) below "
                    f"{MIN_PRACTICE_YD:,} yd:**"
                )
                for w in flagged:
                    yd = _yard_total(w["totals"])
                    st.markdown(
                        f"- _{w['header']}_  — **{yd:,} yd**  "
                        f"(group: {attribution_label(w)}, method: {w['method']})"
                    )
        else:
            st.markdown(
                f"✅ **Week {week_num}** — {wy:,} yd — "
                f"{n_sub} sub-session(s), all ≥ {MIN_PRACTICE_YD:,} yd  "
                f"(min {min_yd:,}, avg {avg_yd:,})"
            )


# ---------------------------------------------------------------------------
# Parse and display.
# ---------------------------------------------------------------------------

with st.spinner("Parsing workouts..."):
    results = compute_workout_totals(text, week_filter=week_filter)

if not results["workouts"]:
    st.warning(
        "No workouts found. The text should contain date headers (e.g. "
        "'Wednesday, August 27, 2025'), 'Coach:' lines, and either "
        "####/#### cumulative checkpoints or NxDIST set lines."
    )
    st.stop()

_render_dedupe_note(results.get("deduped", []))

# Grand total banner.
header_text = f"Grand Total ({results['workouts_parsed']} workout(s))"
if week_filter:
    header_text += f"  —  Week {week_filter} only"
st.subheader(header_text)
render_bucket_row(results["grand_total"])

st.divider()

# Per-week subtotals + drill-in.
st.subheader("Per-Week Subtotals")
for week_num in sorted(results["weekly_subtotals"]):
    wt = results["weekly_subtotals"][week_num]
    wy = _yard_total(wt)
    wm = _meter_total(wt)
    expander_label = f"Week {week_num} — {fmt_amount(wy, wm)}"
    expanded = (week_filter is not None) or (len(results["weekly_subtotals"]) == 1)
    with st.expander(expander_label, expanded=expanded):
        render_bucket_row(wt)
        st.markdown("**Per-day breakdown (mini-microcycle):**")
        st.markdown(daily_to_markdown(results["daily_subtotals"], week_num))
        st.markdown("**Individual workouts:**")
        week_workouts = [w for w in results["workouts"] if w["week"] == week_num]
        st.markdown(workouts_to_markdown(week_workouts))

st.divider()

# Per-week validation pane (per-practice minimum sanity check).
render_weekly_validation(results)

st.divider()

# All workouts table.
st.subheader("All Workouts")
st.markdown(workouts_to_markdown(results["workouts"]))

# Export — both flat CSV (one row per workout) and structured Excel
# (5 sheets: Summary / Per Week / Per Day / Workouts / Validation).
# The Excel mirrors how the page is laid out so the boss can drill
# from the headline down to a single sub-session and the Validation
# sheet lists every flagged practice in one place.
st.subheader("Export")
col_xlsx, col_csv = st.columns(2)
with col_xlsx:
    st.download_button(
        label="Download Excel (5 sheets, week-by-week)",
        data=build_xlsx(results),
        file_name="swim_workouts.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help=(
            "Multi-sheet workbook: Summary, Per Week, Per Day, "
            "Workouts, and Validation. Matches the page layout — "
            "Summary at top, then "
            "drill down by week / day / individual workout."
        ),
    )
with col_csv:
    st.download_button(
        label="Download CSV (flat workouts table)",
        data=workouts_to_csv(results["workouts"]),
        file_name="swim_workouts.csv",
        mime="text/csv",
        help="One row per workout. Useful for pivot-tabling in Excel/Sheets.",
    )
